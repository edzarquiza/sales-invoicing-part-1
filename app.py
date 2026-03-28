import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, datetime
import zipfile
import io
import os

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Dexterous Invoicing Automation",
    page_icon="📄",
    layout="wide",
)

# ── Styles ─────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d6a9f 100%);
        padding: 2rem 2.5rem;
        border-radius: 12px;
        margin-bottom: 2rem;
        color: white;
    }
    .main-header h1 { margin: 0; font-size: 2rem; }
    .main-header p  { margin: 0.4rem 0 0; opacity: 0.85; }
    .step-card {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-left: 4px solid #2d6a9f;
        border-radius: 8px;
        padding: 1.2rem 1.5rem;
        margin-bottom: 1rem;
    }
    .step-card h3 { margin: 0 0 .4rem; color: #1e3a5f; font-size: 1rem; }
    .step-card p  { margin: 0; color: #475569; font-size: .88rem; }
    .metric-card {
        background: white;
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        text-align: center;
    }
    .metric-card .val { font-size: 1.8rem; font-weight: 700; color: #1e3a5f; }
    .metric-card .lbl { font-size: .82rem; color: #64748b; }
    .success-box {
        background: #f0fdf4;
        border: 1px solid #86efac;
        border-radius: 8px;
        padding: 1rem 1.4rem;
        color: #166534;
    }
    .warning-box {
        background: #fffbeb;
        border: 1px solid #fcd34d;
        border-radius: 8px;
        padding: 1rem 1.4rem;
        color: #92400e;
    }
</style>
""", unsafe_allow_html=True)

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1>📄 Dexterous Invoicing Automation</h1>
    <p>Generate per-CM workbooks from XPM data — no more copy-pasting into ChatGPT.</p>
</div>
""", unsafe_allow_html=True)

# ── Helpers ────────────────────────────────────────────────────────────────────

def load_workbook_data(uploaded_file):
    """Load XPM Data and CM assignment from uploaded .xlsm/.xlsx file."""
    file_bytes = uploaded_file.read()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, keep_vba=False)

    # ── XPM Data ──────────────────────────────────────────────────────────────
    ws_xpm = wb["XPM Data"]
    xpm_rows = []
    for row in ws_xpm.iter_rows(values_only=True):
        if all(v is None for v in row):
            break
        xpm_rows.append(row)

    xpm_headers = list(xpm_rows[0])
    xpm_df = pd.DataFrame(xpm_rows[1:], columns=xpm_headers)
    # Keep only the columns we care about (drop the empty 9th col if present)
    cols_needed = ["[Job] Client", "[Staff] Name", "[Time] Date", "[Job] Name",
                   "[Time] Note", "[Time] Billable", "[Time] Time (Totalled)"]
    for c in cols_needed:
        if c not in xpm_df.columns:
            raise ValueError(f"Column '{c}' not found in XPM Data sheet.")
    xpm_df = xpm_df[cols_needed].copy()
    xpm_df["[Time] Date"] = pd.to_datetime(xpm_df["[Time] Date"], errors="coerce")
    xpm_df["[Time] Time (Totalled)"] = pd.to_numeric(
        xpm_df["[Time] Time (Totalled)"], errors="coerce").fillna(0)

    # ── CM assignment ─────────────────────────────────────────────────────────
    ws_cm = wb["CM assignment"]
    cm_rows = []
    for row in ws_cm.iter_rows(values_only=True):
        if all(v is None for v in row):
            break
        cm_rows.append(row)

    cm_headers = list(cm_rows[0])
    cm_df = pd.DataFrame(cm_rows[1:], columns=cm_headers)
    # Expected cols: '[Client] Client', 'CM', 'Batches'
    client_col = cm_df.columns[0]
    cm_col     = cm_df.columns[1]
    batch_col  = cm_df.columns[2]
    cm_df = cm_df[[client_col, cm_col, batch_col]].copy()
    cm_df.columns = ["Client", "CM", "Batch"]
    cm_df["Client"] = cm_df["Client"].astype(str).str.strip()
    cm_df["CM"]     = cm_df["CM"].astype(str).str.strip()
    cm_df["Batch"]  = cm_df["Batch"].astype(str).str.strip()

    wb.close()
    return xpm_df, cm_df


def apply_three_range_filter(xpm_df, cm_df, main_start, main_end, weekly_start, weekly_end, monthly_start, monthly_end):
    """Merge XPM with CM info, then apply the three-range filter logic."""
    # Merge to get Batch per row
    merged = xpm_df.merge(
        cm_df[["Client", "CM", "Batch"]],
        left_on="[Job] Client",
        right_on="Client",
        how="left"
    )
    merged["CM"]    = merged["CM"].fillna("").str.strip()
    merged["Batch"] = merged["Batch"].fillna("").str.strip()

    def in_range(dt, start, end):
        if pd.isnull(dt):
            return False
        return start <= dt.date() <= end

    def row_filter(row):
        b = row["Batch"].lower() if isinstance(row["Batch"], str) else ""
        dt = row["[Time] Date"]
        if b == "weekly":
            return in_range(dt, weekly_start, weekly_end)
        elif b == "monthly":
            return in_range(dt, monthly_start, monthly_end)
        else:
            return in_range(dt, main_start, main_end)

    mask = merged.apply(row_filter, axis=1)
    filtered = merged[mask].copy()
    return filtered


def build_pivot(df):
    """Aggregate by grouping columns, return pivot-style DataFrame."""
    group_cols = ["[Job] Client", "[Staff] Name", "[Time] Date",
                  "[Job] Name", "[Time] Note", "[Time] Billable"]
    agg = (df.groupby(group_cols, dropna=False)["[Time] Time (Totalled)"]
             .sum()
             .reset_index()
             .rename(columns={"[Time] Time (Totalled)": "Sum of [Time] Time (Totalled)"}))
    agg = agg.sort_values(group_cols)
    return agg


HEADER_COLOR = "CCE5FF"
YELLOW_COLOR = "FFFF00"


def write_client_sheet(ws, client_name, client_df):
    """Write a pivot-style sheet for one client."""
    # Two blank rows
    ws.append([])
    ws.append([])

    # Header
    headers = ["Row Labels", "[Staff] Name", "[Time] Date",
               "[Job] Name", "[Time] Note", "[Time] Billable",
               "Sum of [Time] Time (Totalled)"]
    ws.append(headers)
    hdr_row = ws.max_row
    hdr_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    hdr_font = Font(bold=True, name="Arial")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=hdr_row, column=col_idx)
        cell.fill = hdr_fill
        cell.font = hdr_font

    # Data rows with pivot-style blanking
    prev_client = prev_staff = prev_date = prev_job = None
    yellow_fill = PatternFill("solid", fgColor=YELLOW_COLOR)
    bold_font   = Font(bold=True, name="Arial")
    normal_font = Font(name="Arial")

    for _, row in client_df.iterrows():
        client_val = row["[Job] Client"]
        staff_val  = row["[Staff] Name"]
        date_val   = row["[Time] Date"]
        job_val    = row["[Job] Name"]
        note_val   = row["[Time] Note"]
        bill_val   = row["[Time] Billable"]
        hours_val  = row["Sum of [Time] Time (Totalled)"]

        display_client = client_val if client_val != prev_client else None
        display_staff  = staff_val  if staff_val  != prev_staff  else None
        display_date   = date_val   if date_val   != prev_date   else None
        display_job    = job_val    if job_val    != prev_job    else None

        date_str = date_val.strftime("%d/%m/%Y") if pd.notnull(date_val) else ""
        display_date_str = date_str if display_date is not None else None

        ws.append([
            display_client,
            display_staff,
            display_date_str,
            display_job,
            note_val,
            bill_val,
            round(hours_val, 2) if hours_val else 0
        ])

        data_row = ws.max_row
        is_nonbillable = str(bill_val).strip().lower() == "no"

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col_idx)
            if is_nonbillable:
                cell.fill = yellow_fill
            if col_idx in (1, 2):  # Row Labels and Staff Name bold
                cell.font = Font(bold=True, name="Arial",
                                 color="000000" if not is_nonbillable else "000000")
            else:
                cell.font = normal_font

        prev_client = client_val
        prev_staff  = staff_val
        prev_date   = date_val
        prev_job    = job_val

    # Grand Total row
    total = client_df["Sum of [Time] Time (Totalled)"].sum()
    ws.append(["Grand Total", None, None, None, None, None, round(total, 2)])
    gt_row = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=gt_row, column=col_idx)
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.font = Font(bold=True, name="Arial")

    # Auto-fit columns (skip col E = index 5 = [Time] Note)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        if col_idx == 5:  # [Time] Note — keep default
            continue
        max_len = 0
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)


def generate_cm_workbooks(pivot_df, cm_df, main_end_date):
    """Return a dict: {cm_name: {filename: bytes}} and unassigned DataFrame."""
    end_str = main_end_date.strftime("%d.%m.%y")

    # Map client → CM
    client_cm_map = dict(zip(cm_df["Client"], cm_df["CM"]))
    all_clients_in_data = set(pivot_df["[Job] Client"].dropna().unique())

    # Unassigned clients
    unassigned_clients = {c for c in all_clients_in_data
                          if client_cm_map.get(c, "") in ("", "nan", "None")}

    # Group clients by CM
    cm_clients = {}
    for _, row in cm_df.iterrows():
        cm = row["CM"]
        if cm in ("", "nan", "None") or pd.isnull(cm):
            continue
        client = row["Client"]
        if client not in all_clients_in_data:
            continue
        cm_clients.setdefault(cm, []).append(client)

    cm_workbooks = {}   # {cm: bytes}

    for cm, clients in cm_clients.items():
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        for client in clients:
            client_data = pivot_df[pivot_df["[Job] Client"] == client].copy()
            if client_data["Sum of [Time] Time (Totalled)"].sum() == 0:
                continue

            # Truncate sheet name to 31 chars (Excel limit)
            sheet_name = client[:31]
            # Remove invalid characters
            for ch in r"/\?*[]:'":
                sheet_name = sheet_name.replace(ch, " ")
            ws = wb.create_sheet(title=sheet_name)
            write_client_sheet(ws, client, client_data)

        if not wb.sheetnames:
            continue

        buf = io.BytesIO()
        wb.save(buf)
        cm_workbooks[cm] = (f"{cm} {end_str}.xlsx", buf.getvalue())

    # Unassigned workbook
    unassigned_bytes = None
    if unassigned_clients:
        unassigned_df = pivot_df[pivot_df["[Job] Client"].isin(unassigned_clients)]
        wb_u = openpyxl.Workbook()
        wb_u.remove(wb_u.active)
        for client in sorted(unassigned_clients):
            cd = unassigned_df[unassigned_df["[Job] Client"] == client].copy()
            if cd.empty:
                continue
            sheet_name = client[:31]
            for ch in r"/\?*[]:'":
                sheet_name = sheet_name.replace(ch, " ")
            ws = wb_u.create_sheet(title=sheet_name)
            write_client_sheet(ws, client, cd)
        if wb_u.sheetnames:
            buf_u = io.BytesIO()
            wb_u.save(buf_u)
            unassigned_bytes = buf_u.getvalue()

    return cm_workbooks, unassigned_clients, unassigned_bytes


def build_zip(cm_workbooks, unassigned_bytes):
    """Pack all CM workbooks + unassigned into a ZIP, each CM in its own folder."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for cm, (filename, wb_bytes) in cm_workbooks.items():
            zf.writestr(f"{cm}/{filename}", wb_bytes)
        if unassigned_bytes:
            zf.writestr("Unassigned Clients.xlsx", unassigned_bytes)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════

col_steps, col_main = st.columns([1, 2.8], gap="large")

with col_steps:
    st.markdown("""
    <div class="step-card">
        <h3>Step 1 — Upload File</h3>
        <p>Upload your master <code>.xlsm</code> or <code>.xlsx</code> workbook containing <em>XPM Data</em> and <em>CM assignment</em> sheets.</p>
    </div>
    <div class="step-card">
        <h3>Step 2 — Set Date Ranges</h3>
        <p>Enter the three date ranges (Main, Weekly, Monthly) exactly as you would in your ChatGPT prompt.</p>
    </div>
    <div class="step-card">
        <h3>Step 3 — Generate & Download</h3>
        <p>Click <strong>Run Automation</strong> to create per-CM workbooks. Download individually or as a single ZIP.</p>
    </div>
    """, unsafe_allow_html=True)

with col_main:
    # ── Step 1: File upload ───────────────────────────────────────────────────
    st.subheader("① Upload Master Workbook")
    uploaded = st.file_uploader(
        "Drag & drop or click to browse",
        type=["xlsm", "xlsx"],
        help="Must contain sheets: 'XPM Data' and 'CM assignment'",
    )

    if uploaded:
        with st.spinner("Reading workbook…"):
            try:
                xpm_df, cm_df = load_workbook_data(uploaded)
                st.markdown(f"""
                <div class="success-box">
                    ✅ <strong>File loaded successfully.</strong>
                    &nbsp;|&nbsp; XPM rows: <strong>{len(xpm_df):,}</strong>
                    &nbsp;|&nbsp; Clients in CM list: <strong>{len(cm_df):,}</strong>
                </div>
                """, unsafe_allow_html=True)
            except Exception as e:
                st.error(f"❌ Could not load file: {e}")
                st.stop()

        st.markdown("---")
        # ── Step 2: Date ranges ───────────────────────────────────────────────
        st.subheader("② Set Date Ranges")

        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Main Date Range** *(all rows except Weekly / Monthly)*")
            main_start = st.date_input("Main — Start", value=date(2026, 3, 9),  key="ms")
            main_end   = st.date_input("Main — End",   value=date(2026, 3, 22), key="me")
        with c2:
            st.markdown("**Weekly Date Range** *(rows where Batch = \"Weekly\")*")
            weekly_start = st.date_input("Weekly — Start", value=date(2026, 3, 16), key="ws")
            weekly_end   = st.date_input("Weekly — End",   value=date(2026, 3, 22), key="we")

        st.markdown("**Monthly Date Range** *(rows where Batch = \"Monthly\")*")
        m1, m2 = st.columns(2)
        with m1:
            monthly_start = st.date_input("Monthly — Start", value=date(2026, 3, 1),  key="mns")
        with m2:
            monthly_end   = st.date_input("Monthly — End",   value=date(2026, 3, 22), key="mne")

        st.markdown("---")
        # ── Step 3: Run ───────────────────────────────────────────────────────
        st.subheader("③ Generate Workbooks")
        run_btn = st.button("🚀 Run Automation", use_container_width=True, type="primary")

        if run_btn:
            if main_start > main_end or weekly_start > weekly_end or monthly_start > monthly_end:
                st.error("❌ Start date cannot be after end date. Please check your ranges.")
                st.stop()

            with st.spinner("Filtering & pivoting data…"):
                filtered_df = apply_three_range_filter(
                    xpm_df, cm_df,
                    main_start, main_end,
                    weekly_start, weekly_end,
                    monthly_start, monthly_end,
                )
                pivot_df = build_pivot(filtered_df)

            if pivot_df.empty:
                st.markdown("""
                <div class="warning-box">
                    ⚠️ No records matched the given date ranges. 
                    Please double-check the dates and try again.
                </div>
                """, unsafe_allow_html=True)
                st.stop()

            with st.spinner("Building CM workbooks…"):
                cm_workbooks, unassigned_clients, unassigned_bytes = generate_cm_workbooks(
                    pivot_df, cm_df, main_end
                )

            # ── Summary metrics ───────────────────────────────────────────────
            total_rows  = len(filtered_df)
            total_cms   = len(cm_workbooks)
            total_hours = pivot_df["Sum of [Time] Time (Totalled)"].sum()

            mc1, mc2, mc3, mc4 = st.columns(4)
            for col, val, lbl in [
                (mc1, f"{total_rows:,}",        "Filtered Rows"),
                (mc2, f"{total_cms}",            "CM Workbooks"),
                (mc3, f"{total_hours:.1f} hrs",  "Total Hours"),
                (mc4, f"{len(unassigned_clients)}", "Unassigned Clients"),
            ]:
                with col:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="val">{val}</div>
                        <div class="lbl">{lbl}</div>
                    </div>
                    """, unsafe_allow_html=True)

            st.markdown("---")

            # ── Download section ──────────────────────────────────────────────
            st.subheader("⬇️ Download Results")

            # ZIP download (all in one)
            if cm_workbooks:
                zip_bytes = build_zip(cm_workbooks, unassigned_bytes)
                st.download_button(
                    label="📦 Download ALL as ZIP (CM folders + Unassigned)",
                    data=zip_bytes,
                    file_name=f"Invoicing_{main_end.strftime('%d.%m.%y')}.zip",
                    mime="application/zip",
                    use_container_width=True,
                    type="primary",
                )

            st.markdown("**— or download individual CM workbooks —**")

            # Individual downloads in a grid
            cm_list = sorted(cm_workbooks.items())
            for i in range(0, len(cm_list), 3):
                cols = st.columns(3)
                for j, (cm, (filename, wb_bytes)) in enumerate(cm_list[i:i+3]):
                    with cols[j]:
                        st.download_button(
                            label=f"👤 {cm}",
                            data=wb_bytes,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"dl_{cm}",
                        )

            # Unassigned
            if unassigned_bytes:
                st.markdown("---")
                st.markdown(f"""
                <div class="warning-box">
                    ⚠️ <strong>{len(unassigned_clients)} unassigned client(s)</strong> found:
                    {", ".join(sorted(unassigned_clients)[:10])}
                    {"…" if len(unassigned_clients) > 10 else ""}
                </div>
                """, unsafe_allow_html=True)
                st.download_button(
                    label="⚠️ Download Unassigned Clients.xlsx",
                    data=unassigned_bytes,
                    file_name="Unassigned Clients.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
    else:
        st.info("👆 Upload your master workbook above to get started.")

# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    "<p style='text-align:center;color:#94a3b8;font-size:.82rem;'>"
    "Dexterous Invoicing Automation · Built with Streamlit</p>",
    unsafe_allow_html=True,
)
